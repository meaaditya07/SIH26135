"""Sprint 8 — Reporting & exports.

Pure, DB-agnostic helpers that turn tabular data into CSV / XLSX / PDF
byte-streams, plus per-report column definitions and row builders.

The router queries the database and passes a list of dicts here; these
functions do NOT talk to the database so they stay trivially testable.
"""
import csv
import io
from datetime import date, datetime
from typing import Any

_REPORTABLE_TYPES = {
    "scheme-roi",
    "skill-gaps",
    "outcomes",
    "candidates",
    "applications",
}

REPORT_LABELS: dict[str, str] = {
    "scheme-roi": "Scheme ROI",
    "skill-gaps": "Regional Skill Gaps",
    "outcomes": "Employment Outcomes",
    "candidates": "Candidate Registry",
    "applications": "Hiring Pipeline Applications",
}


# ─── Column definitions (report_type -> ordered [(header, key)]) ───

_COLUMNS: dict[str, list[tuple[str, str]]] = {
    "scheme-roi": [
        ("Scheme ID", "scheme_id"),
        ("Period", "period"),
        ("State", "state"),
        ("District", "district"),
        ("Enrolled", "total_enrolled"),
        ("Completed", "total_completed"),
        ("Completion Rate %", "completion_rate"),
        ("Placed 6m", "total_placed_6m"),
        ("Placed 12m", "total_placed_12m"),
        ("Cost / Placement", "cost_per_placement"),
        ("ROI Score", "roi_score"),
        ("Status", "alert_status"),
    ],
    "skill-gaps": [
        ("State", "state"),
        ("District", "district"),
        ("Sector", "sector"),
        ("Skill", "skill_name"),
        ("Demand", "demand_score"),
        ("Supply", "supply_score"),
        ("Gap", "gap_score"),
        ("Direction", "gap_direction"),
    ],
    "outcomes": [
        ("Candidate ID", "candidate_id"),
        ("Interval", "survey_interval"),
        ("Survey Date", "survey_date"),
        ("Employed", "is_employed"),
        ("Self Employed", "is_self_employed"),
        ("Job Title", "current_job_title"),
        ("Salary", "monthly_salary"),
        ("Job Location", "job_location"),
        ("Relevant to Training", "is_job_relevant_to_training"),
    ],
    "candidates": [
        ("Candidate ID", "id"),
        ("Name", "full_name"),
        ("Phone", "phone"),
        ("State", "state"),
        ("District", "district"),
        ("DigiLocker Status", "digilocker_status"),
        ("Skills", "skill_tags"),
        ("Active", "is_active"),
        ("Created", "created_at"),
    ],
    "applications": [
        ("Application ID", "id"),
        ("Job Posting ID", "job_posting_id"),
        ("Candidate ID", "candidate_id"),
        ("Status", "status"),
        ("Match Score", "match_score"),
        ("Applied At", "applied_at"),
        ("Updated At", "updated_at"),
    ],
}


def report_columns(report_type: str) -> list[tuple[str, str]]:
    """Return ordered (display_header, row_key) pairs for a report type."""
    return _COLUMNS[report_type]


def _coerce(value: Any) -> Any:
    """Make a SQLAlchemy/scalar value CSV/Excel-friendly."""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (list, tuple)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return value


def to_csv(report_type: str, rows: list[dict]) -> str:
    """Render rows as a CSV string (header row from the report def)."""
    columns = report_columns(report_type)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([header for header, _ in columns])
    for row in rows:
        writer.writerow([_coerce(row.get(key)) for _, key in columns])
    return buf.getvalue()


def to_xlsx(report_type: str, rows: list[dict]) -> bytes:
    """Render rows as an XLSX workbook in memory and return the bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    columns = report_columns(report_type)
    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_LABELS.get(report_type, report_type)[:31]

    ws.append([header for header, _ in columns])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([_coerce(row.get(key)) for _, key in columns])

    for idx, (header, _) in enumerate(columns, start=1):
        width = min(max(len(header) * 1.6, 12), 45)
        ws.column_dimensions[get_column_letter(idx)].width = width

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def to_pdf(title: str, columns: list[tuple[str, str]], rows: list[dict]) -> bytes:
    """Render rows as a compact PDF table using reportlab (stdlib-free)."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=12 * mm,
        leftMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=14,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        "ReportMeta",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.grey,
        spaceAfter=10,
    )

    story = [
        Paragraph(title, title_style),
        Paragraph(
            f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC · "
            f"{len(rows)} record(s)",
            meta_style,
        ),
        Spacer(1, 2 * mm),
    ]

    headers = [h for h, _ in columns]
    body = [[_coerce(r.get(k)) for _, k in columns] for r in rows]
    table_data = [headers] + body

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(table)
    doc.build(story)
    return buf.getvalue()


def is_reportable(report_type: str) -> bool:
    return report_type in _REPORTABLE_TYPES
